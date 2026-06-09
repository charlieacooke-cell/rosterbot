#!/usr/bin/env python3
"""
Emily's Shift Site Generator
================================
Reads the Paeds ED roster Excel file and generates:
  - emily/index.html  (the "Is Emily Working?" website)
  - emily/shifts.ics  (calendar subscription file)

HOW TO RUN:
  python3 generate_emily_site.py
"""

import json
import uuid
from datetime import datetime, timedelta, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("Missing dependency: run  pip3 install openpyxl  then try again.")

# ============================================================
# CONFIG
# ============================================================

XLSX_FILE    = "emily/FILE_8494.xlsx"
SHEET_NAME   = "Term 2"

OUTPUT_HTML  = "emily/index.html"
OUTPUT_ICS   = "emily/shifts.ics"
SITE_TITLE   = "Is Emily Working?"
PERSON_NAME  = "Emily"
PERSON_EMOJI = "💉"

# ── Shift hours (start, end) decimal 24hr ─────────────────
SHIFT_HOURS = {
    # Term 2 — resident codes
    "D8":    (7.0,  15.5),   # 07:00 – 15:30
    "D10":   (8.0,  18.5),   # 08:00 – 18:30
    "D10WE": (7.0,  17.5),   # 07:00 – 17:30
    "E10":   (14.0, 24.0),   # 14:00 – 00:00
    "E10WE": (14.0, 24.0),   # 14:00 – 00:00
    "E10S":  (11.0, 21.0),   # 11:00 – 21:00  (orientation one-off)
    "P10":   (8.0,  18.5),   # 08:00 – 18:30  (PNW)
    # Previous term — kept so old shifts display correctly
    "wD8":   (8.0,  18.0),
    "N":     (22.5,  8.5),
    "wN":    (22.5,  8.5),
    "E14":   (14.0, 24.0),
    "wE14":  (14.0, 24.0),
    "M13":   (13.0, 23.0),
    "wM13":  (13.0, 23.0),
}

DEFAULT_DAY_HOURS   = (8.0,  18.5)
DEFAULT_NIGHT_HOURS = (22.5,  8.5)

NIGHT_CODES     = {"N", "wN"}
RELIEF_CODES    = {"SR"}
LEAVE_CODES     = {"ADO10", "ADO"}
EVENING_CODES   = {"E10", "E10WE", "E10S"}
AFTERNOON_CODES = {"E14", "wE14", "M13", "wM13"}
DAY_CODES       = {"D8", "D10", "D10WE", "P10", "wD8"}

SHIFT_LABELS = {
    # Term 2
    "D8":    "Day (8 hour)",
    "D10":   "Day (10 hour)",
    "D10WE": "Day (10 hour, weekend)",
    "E10":   "Evening (10 hour)",
    "E10WE": "Evening (10 hour, weekend)",
    "E10S":  "Evening (10 hour)",
    "P10":   "PNW (10 hour)",
    "SR":    "Sick Relief",
    "ADO10": "Rostered Day Off",
    "ADO":   "Rostered Day Off",
    # Previous term
    "wD8":   "Day Shift (Weekend)",
    "N":     "Night Shift",
    "wN":    "Night Shift (Weekend)",
    "E14":   "Afternoon Shift",
    "wE14":  "Afternoon Shift (Weekend)",
    "M13":   "Afternoon Shift",
    "wM13":  "Afternoon Shift (Weekend)",
    # Holiday override
    "AL":    "On Holidays! 🍹",
}

# Codes shown in the edit panel add-shift dropdown
ADDABLE_SHIFTS = [
    ("D8",    "Day (8 hour) – 7:00am to 3:30pm"),
    ("D10",   "Day (10 hour) – 8:00am to 6:30pm"),
    ("D10WE", "Day (10 hour, weekend) – 7:00am to 5:30pm"),
    ("E10",   "Evening (10 hour) – 2:00pm to midnight"),
    ("E10WE", "Evening (10 hour, weekend) – 2:00pm to midnight"),
    ("P10",   "PNW (10 hour) – 8:00am to 6:30pm"),
    ("SR",    "Sick Relief"),
    ("ADO10", "Rostered Day Off"),
]

# ── Hardcoded shifts from previous term ───────────────────
# Recovered from the previous Netlify deploy. Do not edit.
OLD_SHIFTS = {
    "2026-04-27": {"code": "wE14", "type": "afternoon", "label": "Afternoon Shift (Weekend)"},
    "2026-04-28": {"code": "E14",  "type": "afternoon", "label": "Afternoon Shift"},
    "2026-04-29": {"code": "E14",  "type": "afternoon", "label": "Afternoon Shift"},
    "2026-05-04": {"code": "N",    "type": "night",     "label": "Night Shift"},
    "2026-05-05": {"code": "N",    "type": "night",     "label": "Night Shift"},
    "2026-05-06": {"code": "N",    "type": "night",     "label": "Night Shift"},
    "2026-05-10": {"code": "wM13", "type": "afternoon", "label": "Afternoon Shift (Weekend)"},
    "2026-05-11": {"code": "E14",  "type": "afternoon", "label": "Afternoon Shift"},
    "2026-05-12": {"code": "E14",  "type": "afternoon", "label": "Afternoon Shift"},
    "2026-05-13": {"code": "E14",  "type": "afternoon", "label": "Afternoon Shift"},
    "2026-05-18": {"code": "D8",   "type": "day",       "label": "Day Shift"},
    "2026-05-19": {"code": "D8",   "type": "day",       "label": "Day Shift"},
    "2026-05-20": {"code": "N",    "type": "night",     "label": "Night Shift"},
    "2026-05-21": {"code": "N",    "type": "night",     "label": "Night Shift"},
    "2026-05-22": {"code": "N",    "type": "night",     "label": "Night Shift"},
    "2026-05-27": {"code": "D8",   "type": "day",       "label": "Day Shift"},
    "2026-05-28": {"code": "D8",   "type": "day",       "label": "Day Shift"},
    "2026-05-29": {"code": "D8",   "type": "day",       "label": "Day Shift"},
    "2026-05-30": {"code": "wD8",  "type": "day",       "label": "Day Shift (Weekend)"},
    "2026-06-03": {"code": "D8",   "type": "day",       "label": "Day Shift"},
    "2026-06-04": {"code": "D8",   "type": "day",       "label": "Day Shift"},
    "2026-06-05": {"code": "D8",   "type": "day",       "label": "Day Shift"},
    "2026-06-08": {"code": "wN",   "type": "night",     "label": "Night Shift (Weekend)"},
    "2026-06-09": {"code": "N",    "type": "night",     "label": "Night Shift"},
    "2026-06-10": {"code": "N",    "type": "night",     "label": "Night Shift"},
    "2026-06-14": {"code": "wE14", "type": "afternoon", "label": "Afternoon Shift (Weekend)"},
    "2026-06-15": {"code": "E14",  "type": "afternoon", "label": "Afternoon Shift"},
    "2026-06-19": {"code": "D8",   "type": "day",       "label": "Day Shift"},
    "2026-06-20": {"code": "wE14", "type": "afternoon", "label": "Afternoon Shift (Weekend)"},
    "2026-06-21": {"code": "wE14", "type": "afternoon", "label": "Afternoon Shift (Weekend)"},
}

# ── Holiday override ───────────────────────────────────────
HOLIDAY_START = date(2026, 7, 13)
HOLIDAY_END   = date(2026, 7, 26)

# ============================================================
# CORE LOGIC
# ============================================================

def normalize_code(raw):
    """Normalise a raw cell value to a canonical shift code."""
    if not raw:
        return ""
    s = str(raw).strip()
    # Orientation-week special: "E10 (1100 - 2100)" → E10S
    if s.upper().startswith("E10 ("):
        return "E10S"
    # Collapse "D10 WE" / "D10 WE " → "D10WE"
    s = s.replace(" WE", "WE")
    return s

def get_hours(code, shift_type="day"):
    if code in SHIFT_HOURS:
        return SHIFT_HOURS[code]
    return DEFAULT_NIGHT_HOURS if shift_type == "night" else DEFAULT_DAY_HOURS

def decimal_to_hm(h):
    h = h % 24
    return int(h), round((h % 1) * 60)

def classify(code):
    if not code or code in ("None", "0", "1"):
        return "off"
    if code in NIGHT_CODES:
        return "night"
    if code in RELIEF_CODES:
        return "relief"
    if code in LEAVE_CODES:
        return "off"
    if code in EVENING_CODES:
        return "evening"
    if code in AFTERNOON_CODES:
        return "afternoon"
    if code in DAY_CODES:
        return "day"
    return "off"

def label(code):
    return SHIFT_LABELS.get(code, code)

def parse_roster(xlsx_path):
    """
    Parses the Term 2 grid layout.

    The sheet has blocks of 2 weeks: a date row (col B–O contain datetime
    objects), followed by person rows (col A = name, same cols = shift codes).
    We scan linearly: when we hit a date row we cache the dates; when we hit
    Emily's row we map her codes to those dates.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]

    shifts = {}
    current_dates = {}  # col_idx (1-based) → date

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        col_b = row[1].value  # column B (0-based index 1)

        # Date row: col B holds a datetime object
        if isinstance(col_b, (datetime, date)):
            current_dates = {}
            for cell in row[1:15]:  # cols B–O
                val = cell.value
                if isinstance(val, (datetime, date)):
                    d = val.date() if isinstance(val, datetime) else val
                    if d.year < 2026:
                        d = d.replace(year=2026)
                    current_dates[cell.column] = d
            continue

        # Person row: col A must name Emily
        col_a = row[0].value
        if not col_a or not isinstance(col_a, str):
            continue
        name = col_a.strip()
        if "Emily Cooke" not in name or "Reliever for" in name:
            continue

        # Map Emily's shift codes to their dates
        if not current_dates:
            continue
        row_num = row[0].row
        for col_idx, shift_date in current_dates.items():
            raw  = ws.cell(row=row_num, column=col_idx).value
            code = normalize_code(raw)
            if not code or code in ("None", "0", "1"):
                continue
            stype = classify(code)
            if stype == "off":
                continue
            shifts[shift_date.isoformat()] = {
                "code":  code,
                "type":  stype,
                "label": label(code),
            }

    print(f"  ✓ Parsed {len(shifts)} shift entries for {PERSON_NAME}")
    return shifts

# ============================================================
# ICS CALENDAR GENERATOR
# ============================================================

def generate_ics(shifts):
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//IsEmilyWorking//EN",
        f"X-WR-CALNAME:{PERSON_NAME}'s Shifts",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        "REFRESH-INTERVAL;VALUE=DURATION:P1D",
        "X-PUBLISHED-TTL:P1D",
    ]
    for date_str, info in sorted(shifts.items()):
        d     = date.fromisoformat(date_str)
        code  = info["code"]
        stype = info["type"]
        lbl   = info["label"]
        uid   = str(uuid.uuid5(uuid.NAMESPACE_DNS, f"emily-{date_str}"))
        sh, sm = decimal_to_hm(get_hours(code, stype)[0])
        eh, em = decimal_to_hm(get_hours(code, stype)[1])
        if stype == "night":
            dtstart = datetime(d.year, d.month, d.day, sh, sm)
            dtend   = datetime(d.year, d.month, d.day, eh, em) + timedelta(days=1)
            lines += [
                "BEGIN:VEVENT",
                f"UID:{uid}@isemilyworking",
                f"DTSTART:{dtstart.strftime('%Y%m%dT%H%M%S')}",
                f"DTEND:{dtend.strftime('%Y%m%dT%H%M%S')}",
                f"SUMMARY:\U0001f319 {lbl} ({code})",
                f"DESCRIPTION:{PERSON_NAME} is working: {lbl}",
                "END:VEVENT",
            ]
        elif stype in ("day", "afternoon", "evening", "relief"):
            end_h = get_hours(code, stype)[1]
            dtstart = datetime(d.year, d.month, d.day, sh, sm)
            dtend   = (datetime(d.year, d.month, d.day, 0, 0) + timedelta(days=1)
                       if end_h >= 24
                       else datetime(d.year, d.month, d.day, eh, em))
            icons = {"day": "☀️", "afternoon": "🌤", "evening": "🌆", "relief": "\U0001f4df"}
            icon = icons.get(stype, "☀️")
            lines += [
                "BEGIN:VEVENT",
                f"UID:{uid}@isemilyworking",
                f"DTSTART:{dtstart.strftime('%Y%m%dT%H%M%S')}",
                f"DTEND:{dtend.strftime('%Y%m%dT%H%M%S')}",
                f"SUMMARY:{icon} {lbl} ({code})",
                f"DESCRIPTION:{PERSON_NAME} is working: {lbl}",
                "END:VEVENT",
            ]
    lines.append("END:VCALENDAR")
    return "\r\n".join(lines)

# ============================================================
# HTML TEMPLATE
# ============================================================

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>__TITLE__</title>
  <link rel="icon" href="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'><text y='.9em' font-size='90'>__EMOJI__</text></svg>">
  <style>
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
    body {
      font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', sans-serif;
      background: #f0f4f8; color: #1a202c; min-height: 100vh;
    }

    /* ── HEADER ── */
    header {
      background: #fff; border-bottom: 1px solid #e2e8f0;
      padding: 1rem 1.5rem; display: flex; align-items: center; gap: 0.6rem;
    }
    header span { font-size: 1.5rem; }
    header h1 { font-size: 1.2rem; font-weight: 700; color: #2d3748; }
    .edit-trigger {
      margin-left: auto; background: none; border: 1px solid #e2e8f0;
      border-radius: 8px; padding: .3rem .75rem; font-size: .8rem;
      color: #a0aec0; cursor: pointer; flex-shrink: 0; transition: all .15s;
    }
    .edit-trigger:hover { background: #f0fdfa; color: #0d9488; border-color: #99f6e4; }

    /* ── MAIN ── */
    main { max-width: 640px; margin: 0 auto; padding: 1rem 1rem 3rem; }

    /* ── STATUS CARD ── */
    .card {
      background: #fff; border-radius: 16px;
      box-shadow: 0 4px 24px rgba(0,0,0,.08);
      padding: 1.1rem 1.5rem 1rem; text-align: center; margin-bottom: .7rem;
    }
    .today-label { font-size: .75rem; font-weight: 600; letter-spacing: .08em;
                   text-transform: uppercase; color: #718096; margin-bottom: .4rem; }
    .status-emoji  { font-size: 2.8rem; line-height: 1; margin-bottom: .3rem; }
    .status-answer { font-size: clamp(1.7rem, 8vw, 2.4rem); font-weight: 900;
                     line-height: 1.1; margin-bottom: .25rem; }
    .status-shift  { font-size: 1rem; font-weight: 600; margin-bottom: .2rem; }
    .status-sub    { font-size: .85rem; color: #718096; }

    /* status themes */
    .theme-night     { background: #1a237e; color: #fff; }
    .theme-night .status-sub { color: #9fa8da; }
    .theme-day       { background: #fff8e1; color: #92400e; }
    .theme-day .status-sub { color: #b45309; }
    .theme-afternoon { background: #ecfdf5; color: #064e3b; }
    .theme-afternoon .status-sub { color: #059669; }
    .theme-evening   { background: #f5f3ff; color: #4c1d95; }
    .theme-evening .status-sub { color: #7c3aed; }
    .theme-off       { background: #e8f5e9; color: #2e7d32; }
    .theme-off .status-sub { color: #66bb6a; }
    .theme-relief    { background: #fdf6e3; color: #92610a; border: 2px solid #f6d860; }
    .theme-relief .status-sub { color: #b07d2a; }
    .theme-holiday   { background: #fff3e0; color: #e65100; border: 2px solid #ffb74d; }
    .theme-holiday .status-sub { color: #ef6c00; }
    .theme-unknown   { background: #f5f5f5; color: #424242; }

    /* ── NEXT SHIFT ── */
    .next-card {
      background: #fff; border-radius: 12px;
      box-shadow: 0 2px 12px rgba(0,0,0,.06);
      padding: .8rem 1.2rem; margin-bottom: .7rem;
      display: flex; align-items: center; gap: .75rem;
    }
    .next-icon  { font-size: 1.5rem; flex-shrink: 0; }
    .next-title { font-size: .7rem; font-weight: 700; text-transform: uppercase;
                  letter-spacing: .07em; color: #718096; margin-bottom: .1rem; }
    .next-info  { font-size: .95rem; font-weight: 700; color: #2d3748; }
    .next-sub   { font-size: .8rem; color: #718096; }

    /* ── LOOKUP ── */
    .lookup-card {
      background: #fff; border-radius: 12px;
      box-shadow: 0 2px 12px rgba(0,0,0,.06);
      padding: 1rem 1.2rem; margin-bottom: .7rem;
    }
    .lookup-card h2 { font-size: .9rem; font-weight: 700; color: #2d3748; margin-bottom: .7rem; }
    .lookup-row { display: flex; gap: .6rem; }
    .lookup-row input {
      flex: 1; padding: .7rem 1rem; border: 2px solid #e2e8f0;
      border-radius: 10px; font-size: 1rem; outline: none; transition: border-color .2s;
    }
    .lookup-row input:focus { border-color: #0d9488; }
    .lookup-row button {
      padding: .7rem 1.2rem; background: #0d9488; color: #fff;
      border: none; border-radius: 10px; font-size: .95rem;
      font-weight: 600; cursor: pointer; transition: background .2s;
    }
    .lookup-row button:hover { background: #0f766e; }
    #lookup-result { margin-top: 1rem; padding: 1rem; border-radius: 10px;
                     display: none; font-weight: 600; font-size: 1rem; }

    /* ── CALENDAR ── */
    .cal-card {
      background: #fff; border-radius: 12px;
      box-shadow: 0 2px 12px rgba(0,0,0,.06);
      padding: 1rem .8rem; margin-bottom: .7rem;
    }
    .cal-header {
      display: flex; align-items: center; justify-content: space-between;
      padding: 0 .5rem; margin-bottom: 1rem;
    }
    .cal-nav { display: flex; gap: .4rem; }
    .cal-nav button {
      background: #f7fafc; border: 1px solid #e2e8f0; border-radius: 8px;
      padding: .35rem .7rem; cursor: pointer; font-size: 1rem; transition: background .15s;
    }
    .cal-nav button:hover { background: #edf2f7; }
    .cal-month-label { font-size: 1.1rem; font-weight: 800; color: #2d3748;
                       min-width: 160px; text-align: center; }
    .cal-grid { display: grid; grid-template-columns: repeat(7, 1fr); gap: 3px; }
    .cal-day-name { text-align: center; font-size: .7rem; font-weight: 700;
                    text-transform: uppercase; color: #a0aec0; padding: .3rem 0; }
    .cal-day {
      aspect-ratio: 1; display: flex; flex-direction: column;
      align-items: center; justify-content: center;
      border-radius: 8px; cursor: pointer; position: relative;
      transition: transform .1s; font-size: .85rem; font-weight: 500; padding: .1rem;
    }
    .cal-day:hover { transform: scale(1.08); z-index: 1; }
    .cal-day.empty { background: transparent; cursor: default; }
    .cal-day.empty:hover { transform: none; }
    .cal-day.off       { background: #f7fafc; color: #a0aec0; }
    .cal-day.night     { background: #1a237e; color: #fff; }
    .cal-day.day       { background: #fff3e0; color: #c2410c; }
    .cal-day.afternoon { background: #d1fae5; color: #065f46; }
    .cal-day.evening   { background: #ede9fe; color: #5b21b6; }
    .cal-day.relief    { background: #fdf6e3; color: #92610a; border: 1px solid #f6d860; }
    .cal-day.holiday   { background: #ffe0b2; color: #e65100; }
    .cal-day.today-ring    { outline: 3px solid #0d9488; outline-offset: 1px; }
    .cal-day.selected-ring { outline: 3px solid #f6ad55; outline-offset: 1px; }
    .cal-day .day-num  { font-weight: 700; line-height: 1; }
    .cal-day .day-dot  { font-size: .6rem; line-height: 1; }
    .back-to-now-btn {
      margin-top: .8rem; background: rgba(0,0,0,.12); border: 1px solid rgba(0,0,0,.15);
      color: inherit; border-radius: 20px; padding: .3rem .9rem; font-size: .8rem;
      cursor: pointer; font-weight: 600; opacity: .75;
    }
    .back-to-now-btn:hover { opacity: 1; }

    /* ── CAL SUBSCRIPTION ── */
    .sub-card {
      background: #fff; border-radius: 16px;
      box-shadow: 0 2px 12px rgba(0,0,0,.06);
      padding: 1.4rem 1.5rem; margin-bottom: 1.2rem;
    }
    .sub-card h2 { font-size: 1rem; font-weight: 700; color: #2d3748; margin-bottom: .4rem; }
    .sub-card p  { font-size: .9rem; color: #718096; margin-bottom: 1rem; }
    .sub-buttons { display: flex; gap: .6rem; flex-wrap: wrap; }
    .sub-btn {
      padding: .6rem 1rem; border-radius: 10px; font-size: .9rem; font-weight: 600;
      text-decoration: none; display: inline-flex; align-items: center; gap: .4rem;
      transition: opacity .2s;
    }
    .sub-btn:hover { opacity: .85; }
    .sub-btn.google  { background: #4285f4; color: #fff; }
    .sub-btn.apple   { background: #1c1c1e; color: #fff; }
    .sub-btn.outlook { background: #0078d4; color: #fff; }
    .sub-btn.ics     { background: #f0f4f8; color: #4a5568; border: 1px solid #e2e8f0; }

    /* ── LEGEND ── */
    .legend { display: flex; flex-wrap: wrap; gap: .5rem; padding: .2rem .5rem .5rem; }
    .leg-item { display: flex; align-items: center; gap: .4rem; font-size: .8rem; color: #718096; }
    .leg-dot  { width: 12px; height: 12px; border-radius: 3px; flex-shrink: 0; }
    .leg-night     { background: #1a237e; }
    .leg-day       { background: #fff3e0; border: 1px solid #f6ad55; }
    .leg-afternoon { background: #d1fae5; border: 1px solid #34d399; }
    .leg-evening   { background: #ede9fe; border: 1px solid #a78bfa; }
    .leg-relief    { background: #fdf6e3; border: 1px solid #f6d860; }
    .leg-off       { background: #f7fafc; border: 1px solid #e2e8f0; }

    /* ── TOOLTIP ── */
    .tooltip {
      position: fixed; background: #2d3748; color: #fff;
      padding: .5rem .8rem; border-radius: 8px; font-size: .82rem;
      pointer-events: none; z-index: 100; white-space: nowrap;
      transform: translate(-50%, -120%); display: none;
    }

    /* ── FOOTER ── */
    footer { text-align: center; font-size: .78rem; color: #a0aec0; margin-top: 1rem; }
    footer a { color: #a0aec0; }

    /* ── EDIT OVERLAY ── */
    .edit-overlay {
      position: fixed; inset: 0; background: rgba(0,0,0,.55); z-index: 200;
      display: flex; align-items: flex-start; justify-content: center;
      padding: 1.5rem 1rem; overflow-y: auto;
    }
    .edit-panel {
      background: #fff; border-radius: 16px; width: 100%; max-width: 520px;
      padding: 1.5rem; box-shadow: 0 24px 64px rgba(0,0,0,.3); margin-bottom: 1.5rem;
    }
    .edit-panel-hdr {
      display: flex; align-items: center; justify-content: space-between; margin-bottom: 1.2rem;
    }
    .edit-panel-hdr h2 { font-size: 1.1rem; font-weight: 800; color: #2d3748; }
    .close-btn {
      background: none; border: none; font-size: 1.2rem; cursor: pointer;
      color: #a0aec0; padding: .2rem .5rem; border-radius: 6px; line-height: 1;
    }
    .close-btn:hover { background: #f7fafc; color: #2d3748; }
    .pw-label { font-size: .9rem; color: #718096; margin-bottom: .8rem; }
    .pw-row { display: flex; gap: .6rem; }
    .pw-row input {
      flex: 1; padding: .65rem 1rem; border: 2px solid #e2e8f0;
      border-radius: 10px; font-size: 1rem; outline: none; transition: border-color .2s;
    }
    .pw-row input:focus { border-color: #0d9488; }
    .pw-row button {
      padding: .65rem 1.2rem; background: #0d9488; color: #fff;
      border: none; border-radius: 10px; font-size: .95rem; font-weight: 600; cursor: pointer;
    }
    .pw-row button:hover { background: #0f766e; }
    .shift-list {
      max-height: 300px; overflow-y: auto; border: 1px solid #e2e8f0;
      border-radius: 10px; margin-bottom: 1rem;
    }
    .shift-row {
      display: flex; align-items: center; gap: .5rem;
      padding: .5rem .75rem; border-bottom: 1px solid #f0f4f8;
    }
    .shift-row:last-child { border-bottom: none; }
    .shift-date { flex: 2; font-size: .85rem; font-weight: 500; color: #2d3748; }
    .shift-badge {
      font-size: .72rem; font-weight: 700; padding: .15rem .45rem; border-radius: 6px; flex-shrink: 0;
    }
    .badge-night     { background: #1a237e; color: #fff; }
    .badge-day       { background: #fff3e0; color: #c2410c; }
    .badge-afternoon { background: #d1fae5; color: #065f46; }
    .badge-evening   { background: #ede9fe; color: #5b21b6; }
    .badge-relief    { background: #fdf6e3; color: #92610a; }
    .badge-off       { background: #f7fafc; color: #718096; }
    .shift-lbl { flex: 2; font-size: .82rem; color: #718096; }
    .del-btn {
      background: none; border: none; color: #fc8181; cursor: pointer;
      font-size: 1rem; padding: .1rem .4rem; border-radius: 6px; flex-shrink: 0;
    }
    .del-btn:hover { background: #fff5f5; color: #e53e3e; }
    .no-shifts { padding: 1.2rem; color: #a0aec0; text-align: center; font-size: .9rem; }
    .add-section { margin-bottom: 1.1rem; }
    .add-section h3 {
      font-size: .75rem; font-weight: 700; text-transform: uppercase;
      letter-spacing: .07em; color: #718096; margin-bottom: .6rem;
    }
    .add-row { display: flex; gap: .5rem; flex-wrap: wrap; }
    .add-row input[type="date"], .add-row select {
      padding: .6rem .8rem; border: 2px solid #e2e8f0; border-radius: 10px;
      font-size: .9rem; outline: none; background: #fff; transition: border-color .2s;
    }
    .add-row input[type="date"] { flex: 1 1 130px; min-width: 0; }
    .add-row select { flex: 2 1 200px; cursor: pointer; }
    .add-row input:focus, .add-row select:focus { border-color: #0d9488; }
    .add-btn {
      padding: .6rem 1rem; background: #0d9488; color: #fff;
      border: none; border-radius: 10px; font-size: .9rem; font-weight: 600;
      cursor: pointer; flex-shrink: 0;
    }
    .add-btn:hover { background: #0f766e; }
    .edit-actions { display: flex; gap: .6rem; align-items: center; flex-wrap: wrap; }
    .save-btn {
      padding: .7rem 1.4rem; background: #0d9488; color: #fff;
      border: none; border-radius: 10px; font-size: .95rem; font-weight: 700; cursor: pointer;
    }
    .save-btn:hover { background: #0f766e; }
    .save-btn:disabled { opacity: .55; cursor: not-allowed; }
    .cancel-btn {
      padding: .7rem 1.2rem; background: #f7fafc; color: #4a5568;
      border: 1px solid #e2e8f0; border-radius: 10px; font-size: .95rem;
      font-weight: 600; cursor: pointer;
    }
    .cancel-btn:hover { background: #edf2f7; }
    .save-error { font-size: .85rem; color: #e53e3e; display: none; }
  </style>
</head>
<body>
  <header>
    <span>__EMOJI__</span>
    <h1>__TITLE__</h1>
    <button class="edit-trigger" onclick="openEditMode()">Edit</button>
  </header>

  <main>
    <div class="card" id="status-card">
      <div class="today-label" id="today-label">Today</div>
      <div class="status-emoji" id="status-emoji">⏳</div>
      <div class="status-answer" id="status-answer">Loading…</div>
      <div class="status-shift" id="status-shift"></div>
      <div class="status-sub" id="status-sub"></div>
      <button class="back-to-now-btn" id="back-to-now" onclick="resetToNow()" style="display:none">← Back to now</button>
    </div>

    <div class="next-card" id="next-card" style="display:none">
      <div class="next-icon" id="next-icon">📅</div>
      <div>
        <div class="next-title" id="next-title">Next shift</div>
        <div class="next-info" id="next-info"></div>
        <div class="next-sub" id="next-sub"></div>
      </div>
    </div>

    <div class="cal-card">
      <div class="cal-header">
        <div class="cal-nav"><button onclick="changeMonth(-1)">‹</button></div>
        <div class="cal-month-label" id="cal-month-label"></div>
        <div class="cal-nav"><button onclick="changeMonth(1)">›</button></div>
      </div>
      <div class="cal-grid" id="cal-grid"></div>
      <div class="legend" style="margin-top:.8rem">
        <div class="leg-item"><div class="leg-dot leg-night"></div>Night</div>
        <div class="leg-item"><div class="leg-dot leg-day"></div>Day shift</div>
        <div class="leg-item"><div class="leg-dot leg-afternoon"></div>Afternoon shift</div>
        <div class="leg-item"><div class="leg-dot leg-evening"></div>Evening shift</div>
        <div class="leg-item"><div class="leg-dot leg-relief"></div>Sick Relief</div>
        <div class="leg-item"><div class="leg-dot" style="background:#ffe0b2;border:1px solid #ffb74d"></div>Holidays 🍹</div>
      </div>
    </div>

    <div class="lookup-card">
      <h2>🔍 Look up a date</h2>
      <div class="lookup-row">
        <input type="date" id="lookup-input" />
        <button onclick="lookupDate()">Check</button>
      </div>
      <div id="lookup-result"></div>
    </div>

    <div class="sub-card">
      <h2>📅 Add to your calendar</h2>
      <p>Subscribe so Emily's shifts appear automatically in your calendar app.</p>
      <div class="sub-buttons">
        <a class="sub-btn google" id="gcal-btn" href="#" target="_blank">&#x1f4c5; Google Calendar</a>
        <a class="sub-btn apple" id="apple-btn" href="shifts.ics">🍎 Apple Calendar</a>
        <a class="sub-btn outlook" id="outlook-btn" href="#" target="_blank">📧 Outlook</a>
        <a class="sub-btn ics" href="shifts.ics" download>⬇ Download .ics</a>
      </div>
    </div>

    <footer>
      Last updated: <strong>__GENERATED__</strong> ·
      <a href="shifts.ics">shifts.ics</a>
    </footer>
  </main>

  <div class="tooltip" id="tooltip"></div>

  <!-- EDIT OVERLAY -->
  <div class="edit-overlay" id="edit-overlay" style="display:none" onclick="overlayClick(event)">
    <div class="edit-panel">
      <div class="edit-panel-hdr">
        <h2>✏️ Edit Emily's Shifts</h2>
        <button class="close-btn" onclick="closeEditMode()">✕</button>
      </div>
      <div id="pw-section">
        <p class="pw-label">Enter the password to make changes:</p>
        <div class="pw-row">
          <input type="password" id="pw-input" placeholder="Password" autocomplete="current-password" />
          <button onclick="unlockEdit()">Unlock</button>
        </div>
      </div>
      <div id="edit-section" style="display:none">
        <div class="shift-list" id="shift-list"></div>
        <div class="add-section">
          <h3>Add shift</h3>
          <div class="add-row">
            <input type="date" id="add-date" />
            <select id="add-code">__SHIFT_OPTIONS_HTML__</select>
            <button class="add-btn" onclick="addShift()">+ Add</button>
          </div>
        </div>
        <div class="edit-actions">
          <button class="save-btn" id="save-btn" onclick="saveEdits()">Save changes</button>
          <button class="cancel-btn" onclick="closeEditMode()">Cancel</button>
          <span class="save-error" id="save-error"></span>
        </div>
      </div>
    </div>
  </div>

<script>
// ── DATA ─────────────────────────────────────────────────
const BASE_SHIFTS = __SHIFTS_JSON__;
let   SHIFTS      = Object.assign({}, BASE_SHIFTS);
const CODE_INFO   = __CODE_INFO_JSON__;
const HOURS       = __HOURS_JSON__;
const DEFAULT_HOURS = { day: __DEFAULT_DAY_HOURS__, night: __DEFAULT_NIGHT_HOURS__ };

// ── HELPERS ──────────────────────────────────────────────
function toISO(d) {
  return `${d.getFullYear()}-${String(d.getMonth()+1).padStart(2,'0')}-${String(d.getDate()).padStart(2,'0')}`;
}
function friendlyDate(iso) {
  const [y,m,d] = iso.split('-').map(Number);
  return new Date(y,m-1,d).toLocaleDateString('en-GB',{weekday:'long',day:'numeric',month:'long',year:'numeric'});
}
function shiftEmoji(type) {
  return {night:'🌙', day:'☀️', afternoon:'🌤️', evening:'🌆', relief:'📟', off:'💤', holiday:'🍹'}[type] || '❓';
}
function shiftTheme(type) {
  return {
    night:'theme-night', day:'theme-day', afternoon:'theme-afternoon',
    evening:'theme-evening', relief:'theme-relief', off:'theme-off', holiday:'theme-holiday'
  }[type] || 'theme-unknown';
}
function fmtHour(h) {
  if (h === 0 || h === 24) return 'midnight';
  const wh = Math.floor(h % 24), mn = Math.round((h % 1)*60);
  const sfx = wh < 12 ? 'am' : 'pm';
  const h12 = wh === 0 ? 12 : wh <= 12 ? wh : wh - 12;
  return mn > 0 ? `${h12}:${String(mn).padStart(2,'0')}${sfx}` : `${h12}${sfx}`;
}
function getHours(code, type) {
  return HOURS[code] || (type === 'night' ? DEFAULT_HOURS.night : DEFAULT_HOURS.day);
}
function shiftHoursStr(code, type) {
  const [s, e] = getHours(code, type);
  if (type === 'night') return `${fmtHour(s)} – ${fmtHour(e)} next morning`;
  if (e >= 24) return `${fmtHour(s)} – midnight`;
  return `${fmtHour(s)} – ${fmtHour(e)}`;
}
function daysUntil(iso) {
  const today = new Date(); today.setHours(0,0,0,0);
  const [y,m,d] = iso.split('-').map(Number);
  const diff = Math.round((new Date(y,m-1,d) - today) / 86400000);
  return diff === 0 ? 'today' : diff === 1 ? 'tomorrow' : diff === 2 ? 'in 2 days' : `in ${diff} days`;
}
function isDayLike(type) { return type==='day'||type==='afternoon'||type==='evening'; }
function isWorkShift(type) { return isDayLike(type) || type==='night' || type==='relief'; }

// ── STATUS CARD ───────────────────────────────────────────
function setStatus(emoji, answer, shift, sub, type) {
  const card = document.getElementById('status-card');
  card.className = 'card ' + shiftTheme(type);
  document.getElementById('status-emoji').textContent  = emoji;
  document.getElementById('status-answer').textContent = answer;
  document.getElementById('status-shift').textContent  = shift;
  document.getElementById('status-sub').textContent    = sub;
}

function updateStatus() {
  const now      = new Date();
  const hour     = now.getHours() + now.getMinutes() / 60;
  const todayISO = toISO(now);
  const yestISO  = toISO(new Date(now - 86400000));
  document.getElementById('today-label').textContent =
    'Now: ' + now.toLocaleDateString('en-GB',{weekday:'long',day:'numeric',month:'long'});

  // Night from yesterday still running?
  const yInfo = SHIFTS[yestISO];
  if (yInfo && yInfo.type === 'night') {
    const [, yEnd] = getHours(yInfo.code, 'night');
    if (hour < yEnd) {
      setStatus('🌙',yInfo.label,'Emily is at work right now',`until ${fmtHour(yEnd)}`,'night');
      return;
    }
  }

  const info = SHIFTS[todayISO];
  const type = info ? info.type : 'off';

  if (type === 'night') {
    const [nStart, nEnd] = getHours(info.code, 'night');
    if (hour >= nStart) {
      setStatus('🌙',info.label,'Emily is at work right now',`until ${fmtHour(nEnd)} tomorrow`,'night');
    } else if (hour >= 10 && hour < 17) {
      setStatus('😴','Sleeping','Emily should be resting right now',`Night shift starts at ${fmtHour(nStart)}`,'night');
    } else if (hour >= 17) {
      setStatus('🌙','Tonight',`Emily starts work at ${fmtHour(nStart)}`,info.label,'night');
    } else {
      setStatus('🌙','Off','Emily is between night shifts',`Night shift starts at ${fmtHour(nStart)}`,'night');
    }
  } else if (isDayLike(type)) {
    const [dStart, dEnd] = getHours(info.code, 'day');
    const emoji = shiftEmoji(type);
    if (hour < dStart) {
      setStatus(emoji,'Today',`Emily starts work at ${fmtHour(dStart)}`,info.label,type);
    } else if (dEnd < 24 && hour >= dEnd) {
      setStatus('✅','No','Emily has finished for the day',`Shift ended at ${fmtHour(dEnd)}`,'off');
    } else {
      setStatus(emoji,info.label,'Emily is at work right now',
        `until ${dEnd >= 24 ? 'midnight' : fmtHour(dEnd)}`,type);
    }
  } else if (type === 'relief') {
    setStatus('📟','On Call','Available if needed','On call — may be called in to cover','relief');
  } else if (type === 'holiday') {
    setStatus('🍹','On Holidays!','Emily is on annual leave','Enjoy the break 🥂','holiday');
  } else {
    setStatus('✅','No','Emily is not working today','','off');
  }
}

// ── NEXT SHIFT ────────────────────────────────────────────
function updateNextShift() {
  const now      = new Date();
  const hour     = now.getHours() + now.getMinutes() / 60;
  const todayISO = toISO(now);
  const yestISO  = toISO(new Date(now - 86400000));
  const todayInfo = SHIFTS[todayISO];
  const todayType = todayInfo ? todayInfo.type : 'off';
  const yInfo     = SHIFTS[yestISO];
  const overnightActive = yInfo && yInfo.type==='night' && hour < getHours(yInfo.code,'night')[1];
  const baseISO   = (overnightActive && !isWorkShift(todayType)) ? yestISO : todayISO;
  const baseInfo  = SHIFTS[baseISO];
  const baseType  = baseInfo ? baseInfo.type : 'off';
  const midRun    = isWorkShift(baseType);
  const card      = document.getElementById('next-card');

  if (midRun) {
    const workDates = Object.entries(SHIFTS).filter(([,i])=>isWorkShift(i.type)).map(([iso])=>iso).sort();
    const baseIdx = workDates.indexOf(baseISO);
    let end = baseIdx;
    while (end < workDates.length-1) {
      if ((new Date(workDates[end+1]) - new Date(workDates[end])) / 86400000 <= 1) end++; else break;
    }
    let remaining = workDates.slice(baseIdx, end+1);
    if (baseISO===todayISO && isDayLike(todayType)) {
      const [, dEnd] = getHours(todayInfo.code,'day');
      if (dEnd < 24 && hour >= dEnd) remaining = remaining.filter(iso => iso > todayISO);
    }
    if (!remaining.length) return;
    card.style.display = 'flex';
    document.getElementById('next-title').textContent = 'This run';
    document.getElementById('next-icon').textContent  = '📋';
    document.getElementById('next-info').textContent  = `${remaining.length} shift${remaining.length!==1?'s':''} remaining`;
    document.getElementById('next-sub').textContent   = `Run ends ${friendlyDate(workDates[end])}`;
  } else {
    const upcoming = Object.entries(SHIFTS)
      .filter(([iso,i]) => iso > todayISO && isWorkShift(i.type))
      .sort(([a],[b]) => a.localeCompare(b));
    if (!upcoming.length) return;
    const [iso, info] = upcoming[0];
    card.style.display = 'flex';
    document.getElementById('next-title').textContent = 'Next shift';
    document.getElementById('next-icon').textContent  = shiftEmoji(info.type);
    document.getElementById('next-info').textContent  = friendlyDate(iso);
    document.getElementById('next-sub').textContent   = `${info.label} · ${shiftHoursStr(info.code,info.type)} · ${daysUntil(iso)}`;
  }
}

// ── DATE LOOKUP ───────────────────────────────────────────
function lookupDate() {
  const input  = document.getElementById('lookup-input').value;
  const result = document.getElementById('lookup-result');
  if (!input) { result.style.display='none'; return; }
  const info = SHIFTS[input];
  const type = info ? info.type : 'off';
  const hrs  = (info && isWorkShift(type) && type!=='relief') ? ` · ${shiftHoursStr(info.code, type)}` : '';
  const themes = {
    night:     {bg:'#e8eaf6',color:'#1a237e',text:`🌙 Yes — ${info?.label||'Night Shift'}${hrs}`},
    day:       {bg:'#fff3e0',color:'#c2410c',text:`☀️ Yes — ${info?.label||'Day Shift'}${hrs}`},
    afternoon: {bg:'#d1fae5',color:'#065f46',text:`🌤️ Yes — ${info?.label||'Afternoon Shift'}${hrs}`},
    evening:   {bg:'#ede9fe',color:'#5b21b6',text:`🌆 Yes — ${info?.label||'Evening Shift'}${hrs}`},
    relief:    {bg:'#fdf6e3',color:'#92610a',text:`📟 On Call — may be called in`},
    holiday:   {bg:'#fff3e0',color:'#e65100',text:`🍹 On Holidays! Emily is on annual leave`},
    off:       {bg:'#f0fff4',color:'#276749',text:`✅ No — Emily is off`},
  };
  const t = themes[type] || themes.off;
  result.style.display    = 'block';
  result.style.background = t.bg;
  result.style.color      = t.color;
  result.innerHTML        = `<strong>${friendlyDate(input)}</strong><br>${t.text}`;
}
document.getElementById('lookup-input').addEventListener('keydown', e => { if (e.key==='Enter') lookupDate(); });

// ── CALENDAR ──────────────────────────────────────────────
let calYear, calMonth, calSelected = null;
const MONTHS = ['January','February','March','April','May','June',
                'July','August','September','October','November','December'];
const DAYS   = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun'];

function renderCalendar(year, month) {
  calYear = year; calMonth = month;
  document.getElementById('cal-month-label').textContent = `${MONTHS[month]} ${year}`;
  const grid     = document.getElementById('cal-grid');
  const todayISO = toISO(new Date());
  let html = DAYS.map(d=>`<div class="cal-day-name">${d}</div>`).join('');
  const offset = (new Date(year,month,1).getDay()+6)%7;
  for (let i=0;i<offset;i++) html+=`<div class="cal-day empty"></div>`;
  const days = new Date(year,month+1,0).getDate();
  for (let d=1;d<=days;d++) {
    const iso  = `${year}-${String(month+1).padStart(2,'0')}-${String(d).padStart(2,'0')}`;
    const info = SHIFTS[iso];
    const type = info ? info.type : 'off';
    const tip  = info ? info.label : 'No shift';
    const ring = iso===todayISO ? ' today-ring' : (iso===calSelected ? ' selected-ring' : '');
    const dot  = (info&&type!=='off') ? `<div class="day-dot">${shiftEmoji(type)}</div>` : '';
    html += `<div class="cal-day ${type}${ring}" data-iso="${iso}" data-tip="${tip}" onclick="calClick('${iso}')">
               <div class="day-num">${d}</div>${dot}</div>`;
  }
  grid.innerHTML = html;
  grid.querySelectorAll('.cal-day:not(.empty)').forEach(el => {
    el.addEventListener('mousemove', ev => {
      const tt = document.getElementById('tooltip');
      tt.textContent = `${friendlyDate(el.dataset.iso)}: ${el.dataset.tip}`;
      tt.style.display='block'; tt.style.left=ev.clientX+'px'; tt.style.top=ev.clientY+'px';
    });
    el.addEventListener('mouseleave', () => { document.getElementById('tooltip').style.display='none'; });
  });
}

function calClick(iso) {
  const todayISO = toISO(new Date());
  calSelected = iso === todayISO ? null : iso;
  renderCalendar(calYear, calMonth);
  if (iso === todayISO) { resetToNow(); return; }
  showDateStatus(iso);
  document.getElementById('status-card').scrollIntoView({behavior:'smooth',block:'nearest'});
}
function resetToNow() {
  calSelected = null;
  document.getElementById('back-to-now').style.display = 'none';
  updateStatus();
}
function showDateStatus(iso) {
  const info = SHIFTS[iso];
  const type = info ? info.type : 'off';
  document.getElementById('today-label').textContent = friendlyDate(iso);
  document.getElementById('back-to-now').style.display = 'inline-block';
  if (type === 'night') {
    const [nS,nE] = getHours(info.code,'night');
    setStatus('🌙',info.label,`${fmtHour(nS)} – ${fmtHour(nE)} (next day)`,''  ,'night');
  } else if (isDayLike(type)) {
    const [dS,dE] = getHours(info.code,'day');
    setStatus(shiftEmoji(type),info.label,`${fmtHour(dS)} – ${dE>=24?'midnight':fmtHour(dE)}`,'',type);
  } else if (type === 'relief') {
    setStatus('📟','On Call','Available if needed','On call — may be called in','relief');
  } else if (type === 'holiday') {
    setStatus('🍹','On Holidays!','Emily is on annual leave','Enjoy the break 🥂','holiday');
  } else {
    setStatus('✅','Day Off','Not working this day',info?info.label:'','off');
  }
}
function changeMonth(dir) {
  let m=calMonth+dir, y=calYear;
  if (m>11){m=0;y++;} if (m<0){m=11;y--;}
  renderCalendar(y,m);
}

// ── CALENDAR SUBSCRIPTION LINKS ───────────────────────────
function setupCalLinks() {
  const base   = window.location.href.replace(/\/[^/]*$/, '/');
  const icsUrl = base + 'shifts.ics';
  const webcal = icsUrl.replace(/^https?/,'webcal');
  document.getElementById('apple-btn').href   = webcal;
  document.getElementById('gcal-btn').href    = `https://calendar.google.com/calendar/r?cid=${encodeURIComponent(webcal)}`;
  document.getElementById('outlook-btn').href = `https://outlook.live.com/calendar/0/addfromweb?url=${encodeURIComponent(icsUrl)}`;
}

// ── LOAD SERVER OVERRIDES ─────────────────────────────────
async function loadOverrides() {
  try {
    const res = await fetch('/api/shifts');
    if (!res.ok) return;
    const data = await res.json();
    if (data && Object.keys(data).length > 0) {
      SHIFTS = data;
      refresh();
    }
  } catch { /* local file preview — use embedded data */ }
}
function refresh() {
  if (calSelected) showDateStatus(calSelected); else updateStatus();
  updateNextShift();
  renderCalendar(calYear, calMonth);
}

// ── EDIT MODE ─────────────────────────────────────────────
let editShifts = {}, editPassword = '';

function openEditMode() {
  document.getElementById('edit-overlay').style.display = 'flex';
  document.getElementById('pw-input').value = '';
  document.getElementById('pw-section').style.display = 'block';
  document.getElementById('edit-section').style.display = 'none';
  document.getElementById('save-error').style.display = 'none';
  document.getElementById('add-date').value = toISO(new Date());
  setTimeout(() => document.getElementById('pw-input').focus(), 50);
}
function closeEditMode() { document.getElementById('edit-overlay').style.display = 'none'; }
function overlayClick(e) { if (e.target===document.getElementById('edit-overlay')) closeEditMode(); }

function unlockEdit() {
  editPassword = document.getElementById('pw-input').value;
  if (!editPassword) { document.getElementById('pw-input').focus(); return; }
  editShifts = Object.assign({}, SHIFTS);
  document.getElementById('pw-section').style.display = 'none';
  document.getElementById('edit-section').style.display = 'block';
  renderShiftList();
}
function renderShiftList() {
  const dates = Object.keys(editShifts).sort();
  const list  = document.getElementById('shift-list');
  if (!dates.length) {
    list.innerHTML = '<div class="no-shifts">No shifts yet — use the form below to add some.</div>';
    return;
  }
  list.innerHTML = dates.map(iso => {
    const info = editShifts[iso];
    return `<div class="shift-row">
      <span class="shift-date">${friendlyDate(iso)}</span>
      <span class="shift-badge badge-${info.type}">${info.code}</span>
      <span class="shift-lbl">${info.label}</span>
      <button class="del-btn" onclick="deleteShift('${iso}')" title="Delete">✕</button>
    </div>`;
  }).join('');
}
function deleteShift(iso) { delete editShifts[iso]; renderShiftList(); }
function addShift() {
  const iso  = document.getElementById('add-date').value;
  const code = document.getElementById('add-code').value;
  if (!iso) { document.getElementById('add-date').focus(); return; }
  const ci = CODE_INFO[code] || { type: 'day', label: code };
  editShifts[iso] = { code, type: ci.type, label: ci.label };
  renderShiftList();
  const rows = document.getElementById('shift-list').querySelectorAll('.shift-row');
  const idx  = Object.keys(editShifts).sort().indexOf(iso);
  if (rows[idx]) rows[idx].scrollIntoView({ block: 'nearest' });
}
async function saveEdits() {
  const btn = document.getElementById('save-btn');
  const err = document.getElementById('save-error');
  err.style.display = 'none';
  btn.textContent = 'Saving…'; btn.disabled = true;
  try {
    const res = await fetch('/api/shifts', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ password: editPassword, shifts: editShifts }),
    });
    if (res.status === 401) { err.textContent = 'Wrong password.'; err.style.display = 'inline'; }
    else if (!res.ok) { err.textContent = 'Save failed — try again.'; err.style.display = 'inline'; }
    else { SHIFTS = Object.assign({}, editShifts); closeEditMode(); refresh(); }
  } catch { err.textContent = 'Network error — are you online?'; err.style.display = 'inline'; }
  btn.textContent = 'Save changes'; btn.disabled = false;
}

// ── INIT ──────────────────────────────────────────────────
(function init() {
  const now = new Date();
  updateStatus();
  updateNextShift();
  renderCalendar(now.getFullYear(), now.getMonth());
  setupCalLinks();
  loadOverrides();
  document.getElementById('pw-input').addEventListener('keydown', e => {
    if (e.key === 'Enter') unlockEdit();
  });
})();
</script>
</body>
</html>
"""

# ============================================================
# HTML GENERATOR
# ============================================================

def generate_html(shifts, generated_at):
    code_info_json = json.dumps({
        code: {"type": classify(code), "label": label(code)}
        for code, _ in ADDABLE_SHIFTS
    }, separators=(',', ':'))

    shift_options_html = "\n            ".join(
        f'<option value="{code}">{display}</option>'
        for code, display in ADDABLE_SHIFTS
    )

    hours_for_js = {code: list(hrs) for code, hrs in SHIFT_HOURS.items()}

    html = HTML_TEMPLATE
    html = html.replace("__TITLE__",               SITE_TITLE)
    html = html.replace("__EMOJI__",               PERSON_EMOJI)
    html = html.replace("__SHIFTS_JSON__",         json.dumps(shifts, separators=(',', ':')))
    html = html.replace("__CODE_INFO_JSON__",      code_info_json)
    html = html.replace("__SHIFT_OPTIONS_HTML__",  shift_options_html)
    html = html.replace("__GENERATED__",           generated_at)
    html = html.replace("__HOURS_JSON__",          json.dumps(hours_for_js))
    html = html.replace("__DEFAULT_DAY_HOURS__",   json.dumps(list(DEFAULT_DAY_HOURS)))
    html = html.replace("__DEFAULT_NIGHT_HOURS__", json.dumps(list(DEFAULT_NIGHT_HOURS)))
    return html

# ============================================================
# MAIN
# ============================================================

def main():
    script_dir = Path(__file__).parent
    xlsx_path  = script_dir / XLSX_FILE
    html_path  = script_dir / OUTPUT_HTML
    ics_path   = script_dir / OUTPUT_ICS

    print("🏥 Emily's Shift Site Generator")
    print("=" * 40)

    if not xlsx_path.exists():
        desktop = Path.home() / "Desktop" / XLSX_FILE
        if desktop.exists():
            xlsx_path = desktop
            print(f"  (found on Desktop — copy it here for easier future updates)")
        else:
            print(f"\n❌ Excel file not found: {xlsx_path}")
            print(f"   Copy '{XLSX_FILE}' into this folder, then run again.")
            return

    print(f"\n📂 Reading: {xlsx_path.name}")
    shifts = parse_roster(xlsx_path)

    # Merge in historical shifts from previous term (do not overwrite new data)
    for iso, info in OLD_SHIFTS.items():
        if iso not in shifts:
            shifts[iso] = info
    print(f"  ✓ Merged {len(OLD_SHIFTS)} historical shifts")

    # Apply holiday override for Jul 13–26
    d = HOLIDAY_START
    while d <= HOLIDAY_END:
        shifts[d.isoformat()] = {"code": "AL", "type": "holiday", "label": "On Holidays! 🍹"}
        d += timedelta(days=1)
    print(f"  ✓ Applied holiday override ({HOLIDAY_START} – {HOLIDAY_END})")

    print(f"\n  Shift type breakdown:")
    from collections import Counter
    counts = Counter(v["type"] for v in shifts.values())
    for t, n in sorted(counts.items()):
        print(f"    {t}: {n}")

    generated_at = datetime.now().strftime("%d %b %Y at %H:%M")

    html_path.parent.mkdir(exist_ok=True)
    print(f"\n🌐 Writing: {OUTPUT_HTML}")
    html_path.write_text(generate_html(shifts, generated_at), encoding="utf-8")

    print(f"📅 Writing: {OUTPUT_ICS}")
    ics_path.write_text(generate_ics(shifts), encoding="utf-8")

    print(f"\n✅ Done! Generated at {generated_at}")
    print(f"\n   Open {html_path} in a browser to preview.")


if __name__ == "__main__":
    main()
